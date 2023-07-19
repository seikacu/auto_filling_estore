import tkinter as tk

from tkinter import filedialog

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

from openpyxl.utils.exceptions import InvalidFileException

from openpyxl import load_workbook

from auth_data import user, password 


# print("Режимы работы:")
# print("1 - Открыть estore.gz и авторизоваться на сайте")
# print("2 - Ввести ссылку")
# print("3 - Заполнить форму данными из excel (консольный вариант)")
# print("4 - Заполнить форму данными из excel (диалоговое окно)")
# print("5 - Графический режим работы")
#
# mode = int(input("Введите номер режима: "))

options = webdriver.ChromeOptions()

def set_driver_options(options:webdriver.ChromeOptions):
    # user-agent
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 YaBrowser/23.5.2.625 Yowser/2.5 Safari/537.36")

    # for ChromeDriver version 79.0.3945.16 or over
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    options.debugger_address = 'localhost:8989'
    
set_driver_options(options)

caps = DesiredCapabilities().CHROME
caps['pageLoadStrategy'] = 'eager'

service = Service(desired_capabilities=caps, executable_path=r"C:\WebDriver\chromedriver\chromedriver.exe")
driver = webdriver.Chrome(service=service, options=options)

def switch_window(driver:webdriver.Chrome):
    driver.switch_to.window(driver.window_handles[-1])
    driver.execute_script("window.focus();")
    
# switch_window(driver)

try:
    
    # Passing authentication...
    def authentication(driver:webdriver.Chrome):
        try:
            email_input = driver.find_element(By.XPATH, "//input[@name='login[username]']")
            email_input.clear()
            email_input.send_keys(user)

            password_input = driver.find_element(By.ID, "login-password")
            password_input.clear()
            password_input.send_keys(password)
            
            password_input.send_keys(Keys.ENTER)
        except Exception:
            print("Поля аутентификации не найдены или Вы уже авторизованы")
            pass

    # заполнить set_textarea на сайте
    def set_textarea(driver:webdriver.Chrome, name, arg, j):
        try:
            textarea = driver.find_element(By.XPATH, f"//textarea[@name='offer[offerRows][items][{j}][{arg}]']")
            textarea.clear()
            textarea.send_keys(name)
        except NoSuchElementException:
            print(f"Поле {arg} не найдено")
            pass

    # заполнить input на сайте
    def set_input(driver:webdriver.Chrome, name, arg, j):
        try:
            inputVal = driver.find_element(By.XPATH, f"//input[@name='offer[offerRows][items][{j}][{arg}]']")
            inputVal.clear()
            inputVal.send_keys(name)
        except NoSuchElementException:
            print(f"Поле {arg} не найдено")
            pass
    
    # сохранить введенные данные в форме 
    def save_form(driver:webdriver.Chrome):
        try:
            button = driver.find_element(By.XPATH, f"//input[@value='Сохранить черновик оферты']")
            button.click()
        except NoSuchElementException:
            print(f"Кнопка [Сохранить черновик оферты] не найдена")
            pass        
       
    def get_wb():
        
        fileName = str(input("Введите имя файла excel (без .xlsx): "))
        # Load workbook
        wb = load_workbook(f"./{fileName}.xlsx")
        return wb
        
    def get_sheet(wb:load_workbook):
               
        # Get a sheet by name 
        sheet = wb['Лист1']
        return sheet
    
    # заполнение формы данными из excel (диалоговый режим)
    def process_wb(wb:load_workbook):
        
        sheet = get_sheet(wb)
        fill_form_from_sheet(sheet)
        # save_form(driver)
        wb.close()
    
    # открыть файл excel через диалоговое окно
    def open_xlsx_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            try:
                wb = load_workbook(file_path)
                print("Файл успешно открыт:", file_path)
                # заполнение формы данными из excel
                process_wb(wb)
            except InvalidFileException:
                print("Ошибка: Неверный формат файла или файл поврежден.")
            except Exception as e:
                print("Произошла ошибка:", e)
    
    # заполнение формы данными из excel
    def fill_form_from_sheet(sheet):
        i = 2
        j = 0
        while True:
            
            # Наименование товара, работы, услуги *
            nameProduct = sheet[f'A{i}'].value 
            if nameProduct == None:
                break
            
            # Уточняющие характеристики
            characteristics = sheet[f'B{i}'].value 
            # Срок гарантии *
            warranty = sheet[f'C{i}'].value 
            # Цена без НДС за единицу (руб.) *
            # PriceWithoutVAT = sheet[f'D{i}'].value 
            # Цена с НДС за единицу (руб.) *
            PriceWithVAT = sheet[f'E{i}'].value 
            # Ставка НДС (%): *
            # VATRate = sheet[f'F{i}'].value
                            
            # заполнить "Уточняющие характеристики" на сайте
            set_textarea(driver, characteristics, "comment", j)
            # заполнить "Уточняющие характеристики" на сайте
            set_textarea(driver, warranty, "warranty", j)
            # заполнить "Цена с НДС за единицу (руб.) *" на сайте
            set_input(driver, str(PriceWithVAT), "offerPriceNdsView", j)
            
            i += 1
            j += 1
        

    def on_ctrl_v():
        text = root.clipboard_get()
        root.event_generate("<<Paste>>", data=text)
            
    def on_button1_click():
        print("Запущена авторизация")
        switch_window(driver)
        driver.get("https://account.gz-spb.ru/login")
        authentication(driver)
    
    def on_button2_click():
        print("Запущен переход по ссылке")
        switch_window(driver)
        def handle_text():
            text = text_input.get("1.0", "end-1c")
            if text.strip():
                print("Введенный текст:", text)
                driver.get(text)
                text_window.destroy()  # Закрываем окно после обработки текста
                root.deiconify()  # Восстанавливаем видимость главного окна
                
        def paste_text(event):
            text = root.clipboard_get()
            text_input.insert("insert", text)
            return "break"  # Останавливаем дальнейшую обработку событий
        
        def on_closing():
            root.deiconify()  # Восстанавливаем видимость главного окна при закрытии диалогового окна
            text_window.destroy()

        # Скрываем главное окно перед созданием диалогового окна
        root.withdraw()
    
        # Создаем окно с вводом текста
        text_window = tk.Toplevel(root)
        text_window.title("Введите ссылку для перехода")

        # Обработчик закрытия окна по крестику
        text_window.protocol("WM_DELETE_WINDOW", on_closing)
    
        # Создаем виджет Text в окне
        text_input = tk.Text(text_window, height=3, width=80)
        text_input.pack(pady=10)
        
        # Привязываем обработчик вставки текста по Ctrl+V
        text_input.bind("<Control-v>", paste_text)
    
        # Создаем кнопку обработки текста в окне
        process_button = tk.Button(text_window, text="Ок", command=handle_text)
        process_button.pack(pady=10)
                
    def on_button3_click():
        print("Запущен ввод excel файла")
        switch_window(driver)
        open_xlsx_file()

    
    # # режим 1
    # if mode == 1:
    #     switch_window(driver)
    #     driver.get("https://account.gz-spb.ru/login")
    #     authentication(driver)
    #
    # # режим 2
    # if mode == 2:
    #     switch_window(driver)
    #     url = str(input("Введите ссылку: "))
    #     driver.get(url)
    #
    # # режим 3
    # if mode == 3:
    #     switch_window(driver)
    #     wb = get_wb()
    #     sheet = get_sheet(wb)
    #     fill_form_from_sheet(sheet)
    #     save_form(driver)
    #     wb.close()
    #
    # # режим 4
    # if mode == 4:
    #     switch_window(driver)
    #     # Создание окна
    #     root = tk.Tk()
    #     root.title("Открыть XLSX файл")
    #
    #     # Создание кнопки "Открыть файл"
    #     open_button = tk.Button(root, text="Открыть XLSX файл", command=open_xlsx_file)
    #     open_button.pack(pady=20)
    #
    #     # Привязка обработчика событий для Ctrl+V
    #     root.bind("<Control-v>", on_ctrl_v)
    #
    #     # Запуск основного цикла обработки событий
    #     root.mainloop()
    #
    # # режим 5 графический
    # if mode == 5:
    # switch_window(driver)
    
    # Создание окна
    root = tk.Tk()
    root.title("АИС ГЗ: заполнение формы")
                
    # Создание кнопки 1
    button1 = tk.Button(root, text="Авторизоваться", command=on_button1_click)
    button1.pack(side=tk.LEFT, padx=15, pady=15)
        
    # Создание кнопки 2
    button2 = tk.Button(root, text="Ввести ссылку", command=on_button2_click)
    button2.pack(side=tk.LEFT, padx=15, pady=15)
        
    # Создание кнопки 3
    button3 = tk.Button(root, text="Выбрать и загрузить файл импорта excel", command=on_button3_click)
    button3.pack(side=tk.LEFT, padx=15, pady=15)
                
    # Установка окна "always on top" (всегда сверху)
    root.wm_attributes("-topmost", 1)

    # Запуск основного цикла обработки событий
    root.mainloop()
        
    print("Программа завершена")
        
except Exception as ex:
    print(ex)
finally:
    # driver.close()
    driver.quit()